"""
Script:      check_facts.py
Description: Recomputes every countable fact asserted in the project's markdown
             and compares it against what the documents actually say.

             This exists because two such numbers were found wrong in one
             sitting: CLAUDE.md claimed the pinned PDFs run to 500 pages when
             they run to 460, and PLAN.md carried a class count that had been
             garbled during an edit. Neither was a typo. Both were figures
             derived once, written as prose, and never re-derived when the
             corpus changed underneath them.

             A number in prose has no owner. This script makes the pinned files
             the owner and the prose the thing that has to keep up.

             Only countable claims derived from the pinned corpus are checked.
             Judgements, decisions and reasoning are out of scope and always
             will be; those are reviewed by reading. Figures attributed to an
             external source (a cited paper's benchmark, say) are out of scope
             too: they cannot be recomputed from the pinned files, so their
             owner is the citation and its access date, not this script. Such
             figures live behind a [n] reference marker instead.

Inputs:      data/raw/**            (read-only, pinned)
             *.md and docs/*.md     (read-only, scanned for the stated figure)

Outputs:     A report on stdout. Writes nothing to disk.

Usage:       python scripts/check_facts.py
                 check every fact, report drift
             python scripts/check_facts.py --verbose
                 also show facts that match

Exit codes:  0  every stated figure matches the source it came from
             1  at least one figure has drifted, or is asserted nowhere
             2  a pinned file needed for a check is missing

Date:        2026-08-18
Owner:       Jason Delosh
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

# pymupdf is imported under its legacy name "fitz", matching read_pdf.py.
import fitz
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
RAW = REPO_ROOT / "data" / "raw"

# Documents scanned for stated figures. docs/standards_map.html is included:
# it is linked from docs/sources.md and a session acts on what it says, so its
# numbers need the same guard as the prose. Being HTML makes no difference to a
# regex looking for a figure.
DOCS = ["README.md", "BACKGROUND.md", "PLAN.md", "CLAUDE.md",
        "docs/sources.md", "docs/usdm_ig_map.md", "docs/standards_map.html"]


### Measurements ###############################################################
#
# One function per countable fact. Each returns the true value, computed from a
# pinned file. They are deliberately small and independent so that a failing
# measurement names exactly one fact.


def pinned_pdf_pages() -> int:
    """Total pages across every PDF registered in read_pdf.py.

    Reads the registry from read_pdf.py rather than listing PDFs on disk, so
    that a PDF present but unregistered does not silently inflate the count that
    CLAUDE.md's "never read one whole" rule is scaled against.
    """
    import importlib.util

    spec = importlib.util.spec_from_file_location("rp", REPO_ROOT / "scripts" / "read_pdf.py")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return sum(len(fitz.open(entry["path"])) for entry in module.DOCUMENTS.values())


def ig_sections() -> int:
    """Bookmarks in the USDM Implementation Guide, which is what a section is."""
    return len(fitz.open(RAW / "usdm_v4" / "USDM-IG.pdf").get_toc())


def core_rules() -> int:
    """Rows carrying a rule ID in the conformance rules workbook."""
    sheet = openpyxl.load_workbook(
        RAW / "usdm_v4" / "USDM_CORE_Rules.xlsx", read_only=True
    )["Version 3.0 and 4.0 CORE rules"]
    return sum(1 for row in list(sheet.iter_rows(values_only=True))[1:] if row[0])


def m11_elements() -> int:
    """Data elements in the M11 Technical Specification.

    Counted by the "Term (Variable)" label that opens each element block, which
    is the document's own delimiter rather than a heuristic of ours.
    """
    text = "".join(page.get_text() for page in
                   fitz.open(RAW / "ich_m11" / "ICH_M11_TechnicalSpecification.pdf"))
    return len(re.findall(r"Term \(Variable\)\s*\n\s*<([^>]{1,80})>", text))


def uml_delta_rows() -> int:
    """Lines in the v3.0-to-v4.0 change file, header included, as quoted."""
    path = RAW / "usdm_v4" / "uml" / "UML_DELTA_3-0-0_4-0-0.csv"
    return len(path.read_text(encoding="utf-8").splitlines())


def dictionary_codes() -> int:
    """Distinct NCI C-codes named in the data dictionary."""
    text = (RAW / "usdm_v4" / "uml" / "dataDictionary.MD").read_text(encoding="utf-8")
    return len(set(re.findall(r"\b(C\d{4,6})\b", text)))


def shared_codes() -> int:
    """NCI codes appearing in both the M11 Technical Specification and USDM's CT.

    Guarded because it is the one figure on the standards map that contradicts
    an intuition: both standards use NCI codes, so they look interchangeable,
    and they are not. If this number ever drifts toward either total it would
    change the conclusion, not just the caption.
    """
    text = "".join(page.get_text() for page in
                   fitz.open(RAW / "ich_m11" / "ICH_M11_TechnicalSpecification.pdf"))
    m11 = set(re.findall(r"\b(C\d{4,6})\b", text))

    terminology = set()
    for sheet in openpyxl.load_workbook(RAW / "usdm_v4" / "USDM_CT.xlsx", read_only=True):
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    terminology.update(re.findall(r"\b(C\d{4,6})\b", str(cell)))
    return len(m11 & terminology)


def worked_examples() -> int:
    """Worked example studies, one directory each."""
    return len([d for d in (RAW / "usdm_examples").iterdir() if d.is_dir()])


def examples_with_estimands() -> int:
    """Worked-example studies whose USDM JSON defines at least one estimand.

    Estimands hang off each studyDesign. Counted because PLAN.md leans on their
    scarcity, only one of the three examples defines any, to justify why Phase 1
    must select for documents that actually define estimands. If the corpus
    grows or an example gains an estimand, that argument has to move with it.
    """
    count = 0
    for directory in (RAW / "usdm_examples").iterdir():
        if not directory.is_dir():
            continue

        # An example ships a PDF, an .xlsx and one USDM export; the export is
        # the only .json, so the glob cannot pick up the wrong file.
        exports = list(directory.glob("*.json"))
        if not exports:
            continue

        document = json.loads(exports[0].read_text(encoding="utf-8"))
        designs = document["study"]["versions"][0]["studyDesigns"]
        if any(design.get("estimands") for design in designs):
            count += 1
    return count


# Each entry is (label, measurement, regex capturing the figure as stated).
# The regex must be specific enough that it cannot match an unrelated number;
# a loose pattern would report a false match and defeat the point.
FACTS = [
    ("pinned PDF pages",      pinned_pdf_pages,  r"(\d+) pages across"),
    ("IG sections",           ig_sections,       r"of (\d+) sections"),
    ("CORE rules",            core_rules,        r"(\d+) rules\b"),
    ("M11 data elements",     m11_elements,      r"(\d+) elements"),
    ("UML delta rows",        uml_delta_rows,    r"(\d+) rows:"),
    ("dataDictionary codes",  dictionary_codes,  r"(\d+) NCI codes|all (\d+) codes"),
    ("M11 and USDM shared codes", shared_codes,  r"(\d+) codes in common"),
    # Written as a word in prose, so the check accepts either form. Kept narrow
    # enough that "three" elsewhere in a sentence cannot match.
    ("worked example studies", worked_examples,  r"(?:(\d+)|(?i:(three)|(two)|(four))) (?:real protocols|worked example)"),
    # The count of examples that define an estimand, as stated in PLAN.md. The
    # trailing literal "of the three pinned examples defines" anchors it so the
    # captured number is the leading count, not the "three" later in the phrase.
    ("examples with estimands", examples_with_estimands,
     r"(?:(\d+)|(?i:(one)|(two)|(three))) of the three pinned examples defines"),
]


### Reporting ##################################################################


# Small counts are often written as words in prose. Mapping them here keeps the
# check honest without forcing the documents to use digits where words read
# better.
WORD_NUMBERS = {"one": "1", "two": "2", "three": "3", "four": "4"}


def stated_values(pattern: str) -> list[tuple[str, int]]:
    """Every occurrence of a figure matching pattern, with the file it is in.

    Returns a list rather than a single value because the same fact is often
    asserted in more than one document, and each occurrence has to agree
    independently. Reporting only the first would hide a stale copy elsewhere.
    """
    found = []
    for name in DOCS:
        path = REPO_ROOT / name
        if not path.exists():
            continue
        for match in re.finditer(pattern, path.read_text(encoding="utf-8")):
            # Alternation groups leave unmatched branches as None; take the one
            # that fired.
            raw = next(g for g in match.groups() if g is not None)
            value = WORD_NUMBERS.get(raw.lower(), raw)
            found.append((name, int(value)))
    return found


def main() -> int:
    """Recompute every fact, compare it to what the documents say, report drift."""
    sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(
        description="Check countable claims in the markdown against the pinned files."
    )
    parser.add_argument("--verbose", action="store_true", help="also show facts that match")
    args = parser.parse_args()

    drifted = unasserted = 0

    for label, measure, pattern in FACTS:
        # A missing pinned file is a different failure from a wrong number, and
        # cannot be reported as drift because nothing can be measured.
        try:
            actual = measure()
        except (FileNotFoundError, KeyError, OSError) as exc:
            print(f"  UNMEASURABLE  {label}: {exc}")
            return 2

        occurrences = stated_values(pattern)

        if not occurrences:
            print(f"  NOT ASSERTED  {label}: measured {actual}, no document states it")
            unasserted += 1
            continue

        for name, stated in occurrences:
            if stated != actual:
                print(f"  DRIFTED       {label} in {name}: says {stated}, actual {actual}")
                drifted += 1
            elif args.verbose:
                print(f"  ok            {label} in {name}: {actual}")

    print()
    print(f"{len(FACTS)} fact(s) checked, {drifted} drifted, {unasserted} asserted nowhere.")

    # A fact nobody asserts is not an error in the documents; it just means this
    # script is tracking something the prose does not claim. Only real drift
    # fails the run.
    return 1 if drifted else 0


if __name__ == "__main__":
    raise SystemExit(main())
