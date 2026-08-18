"""
Script:      read_xlsx.py
Description: Reads a sheet out of any of the project's pinned Excel workbooks and
             prints it as text, so a working session can consult controlled
             terminology, conformance rules, and the worked-example source
             spreadsheets without opening Excel.

             Four kinds of workbook are in scope and they differ wildly in shape:
             USDM_CT.xlsx has 2 sheets, USDM_CORE_Rules.xlsx has 1, and each
             worked example has 25 to 35. The example workbooks include the
             mainTimeline sheet, which is the Schedule of Activities grid and can
             run to 58 columns. A fixed table layout is unreadable at that width,
             so --format records prints one field per line instead.

Inputs:      Any .xlsx under data/. Opened read-only; nothing is written back.
Outputs:     Plain text on stdout. Writes nothing to disk.

Usage:       python scripts/read_xlsx.py <workbook>
                 list the sheets, with row and column counts
             python scripts/read_xlsx.py <workbook> --sheet mainTimeline
                 print one sheet as an aligned table
             python scripts/read_xlsx.py <workbook> --sheet study --format records
                 print one sheet one field per line, for wide sheets
             python scripts/read_xlsx.py <workbook> --find "Screening"
                 search every sheet for a term
             python scripts/read_xlsx.py --all --find "epoch"
                 search every workbook under data/

Exit codes:  0 success
             1 the workbook was not found, or the named sheet does not exist

Date:        2026-08-17
Author:      Jason Delosh
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

### Constants ##################################################################

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data"

# Excel writes a "~$name.xlsx" lock file beside any workbook that is currently
# open. It is not a real workbook and openpyxl raises PermissionError on it, so
# it is filtered out of every discovery path rather than handled at open time.
LOCK_FILE_PREFIX = "~$"

# Width beyond which a cell's text is truncated in table format. Chosen so that
# a sheet of ~8 columns still fits a normal terminal; wider sheets should use
# --format records instead of being squeezed further.
MAX_CELL_WIDTH = 40


### Discovery ##################################################################


def find_workbooks() -> list[Path]:
    """
    Return every real .xlsx under data/, sorted, excluding Excel lock files.

    Used by --all. Sorted so that repeated runs list workbooks in the same order
    and output can be diffed between sessions.
    """
    return sorted(
        path
        for path in DATA_DIR.rglob("*.xlsx")
        if not path.name.startswith(LOCK_FILE_PREFIX)
    )


def resolve_workbook(argument: str) -> Path | None:
    """
    Turn a user-supplied workbook argument into a real path.

    Accepts a full path, a path relative to the repo root, or just a filename,
    because typing the full path to a nested example workbook is tedious. A bare
    filename is matched case-insensitively against every workbook under data/,
    and a partial name is accepted if it matches exactly one workbook.

    Returns None when nothing matches or when a partial name is ambiguous; the
    caller reports the failure.
    """
    direct = Path(argument)
    if direct.is_file():
        return direct

    relative = REPO_ROOT / argument
    if relative.is_file():
        return relative

    needle = argument.lower()
    matches = [p for p in find_workbooks() if needle in p.name.lower()]

    if len(matches) == 1:
        return matches[0]

    if len(matches) > 1:
        print(f"{argument!r} matches {len(matches)} workbooks:", file=sys.stderr)
        for path in matches:
            print(f"  {path.relative_to(REPO_ROOT)}", file=sys.stderr)

    return None


### Cell handling ##############################################################


def cell_text(value) -> str:
    """
    Render one cell value as a single-line string.

    Empty cells come back from openpyxl as None and are rendered as an empty
    string rather than the literal "None", which would otherwise fill the sparse
    Schedule of Activities grid with noise.

    Newlines inside a cell are replaced rather than kept, because a multi-line
    cell would break row alignment in table format. Protocol spreadsheets do use
    multi-line cells for footnote text, so this is not a rare case.
    """
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def read_rows(worksheet) -> list[list[str]]:
    """
    Read a worksheet into a list of string rows, dropping fully empty rows.

    Trailing empty rows are common in these workbooks because openpyxl reports
    max_row from the sheet dimensions, which often overshoot the real data.
    Dropping empty rows keeps the output honest about how much content there is.
    """
    rows = []
    for raw_row in worksheet.iter_rows(values_only=True):
        cells = [cell_text(value) for value in raw_row]
        if any(cells):
            rows.append(cells)
    return rows


### Output formats #############################################################


def print_table(rows: list[list[str]]) -> None:
    """
    Print rows as a column-aligned table with the first row treated as a header.

    Column widths are computed from the content so narrow columns stay narrow.
    Cells longer than MAX_CELL_WIDTH are truncated with an ellipsis; the full
    value is still available via --format records, so nothing is unrecoverable.
    """
    if not rows:
        print("(sheet is empty)")
        return

    trimmed = [
        [
            cell if len(cell) <= MAX_CELL_WIDTH else cell[: MAX_CELL_WIDTH - 3] + "..."
            for cell in row
        ]
        for row in rows
    ]

    column_count = max(len(row) for row in trimmed)
    widths = [0] * column_count
    for row in trimmed:
        for index, cell in enumerate(row):
            widths[index] = max(widths[index], len(cell))

    for row_number, row in enumerate(trimmed):
        padded = row + [""] * (column_count - len(row))
        print("  ".join(cell.ljust(widths[i]) for i, cell in enumerate(padded)).rstrip())

        # A rule under the header row, so the header is visually distinct without
        # needing colour, which does not survive being piped or pasted.
        if row_number == 0:
            print("  ".join("-" * width for width in widths).rstrip())


def print_records(rows: list[list[str]]) -> None:
    """
    Print each data row as a block of "header: value" lines.

    This is the format for wide sheets. The Schedule of Activities grid runs to
    58 columns, where a table is unreadable and truncation would hide the visit
    column names that matter. Empty fields are skipped, which matters because
    an SoA grid is mostly empty by design.
    """
    if not rows:
        print("(sheet is empty)")
        return

    headers = rows[0]

    for row_number, row in enumerate(rows[1:], start=2):
        print(f"--- row {row_number} ---")
        for index, cell in enumerate(row):
            if not cell:
                continue
            header = headers[index] if index < len(headers) else f"col{index + 1}"
            print(f"  {header or f'col{index + 1}'}: {cell}")
        print()


### Search #####################################################################


def search_workbook(path: Path, term: str) -> list[str]:
    """
    Find every cell in every sheet of one workbook containing term.

    Reports sheet name, row number, and the full cell text. Row numbers are
    1-indexed to match what Excel shows, so a hit can be looked up by hand.

    Returns a list of formatted lines; printing is left to the caller so that
    --all can group output per workbook.
    """
    needle = term.lower()
    hits = []

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_number, raw_row in enumerate(
                worksheet.iter_rows(values_only=True), start=1
            ):
                for value in raw_row:
                    text = cell_text(value)
                    if needle in text.lower():
                        if len(text) > 120:
                            text = text[:120] + "..."
                        hits.append(f"  {sheet_name} row {row_number}: {text}")
                        # One hit per row is enough to locate it; a row where the
                        # term appears in several cells would otherwise repeat.
                        break
    finally:
        # read_only mode holds the file handle open until closed explicitly,
        # which on Windows blocks anything else from opening the workbook.
        workbook.close()

    return hits


### Entry point ################################################################


def main() -> int:
    """
    Parse arguments, dispatch to one mode, and return a shell exit code.

    Modes are checked in order of specificity: --all --find searches every
    workbook, --find searches one, --sheet prints one sheet, and a bare workbook
    name lists its sheets.
    """
    # openpyxl returns proper Unicode, but a Windows console defaults to a
    # legacy code page and silently replaces anything it cannot encode. Controlled
    # terminology and rule text use em dashes, bullets and non-ASCII quotes, so
    # without this the output is mangled exactly where it carries meaning.
    sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(
        description="Read a sheet from one of the project's pinned Excel workbooks."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        help="path, filename, or unique fragment of a filename under data/",
    )
    parser.add_argument("--sheet", help="sheet to print; omit to list sheets")
    parser.add_argument("--find", help="search every sheet for a term")
    parser.add_argument(
        "--all",
        action="store_true",
        help="apply --find across every workbook under data/",
    )
    parser.add_argument(
        "--format",
        choices=("table", "records"),
        default="table",
        help="table (default) or records, one field per line, for wide sheets",
    )
    args = parser.parse_args()

    # Mode: search every workbook. Handled before workbook resolution because
    # --all makes the positional workbook argument meaningless.
    if args.all:
        if not args.find:
            print("--all requires --find", file=sys.stderr)
            return 1
        for path in find_workbooks():
            hits = search_workbook(path, args.find)
            if hits:
                print(f"=== {path.relative_to(REPO_ROOT)}  ({len(hits)} hit(s))")
                print("\n".join(hits))
                print()
        return 0

    if not args.workbook:
        print("Workbooks under data/:\n", file=sys.stderr)
        for path in find_workbooks():
            print(f"  {path.relative_to(REPO_ROOT)}", file=sys.stderr)
        return 1

    path = resolve_workbook(args.workbook)
    if path is None:
        print(f"No workbook matching {args.workbook!r}.", file=sys.stderr)
        return 1

    # Mode: search one workbook.
    if args.find:
        hits = search_workbook(path, args.find)
        if not hits:
            print(f"No cells contain {args.find!r} in {path.name}.")
            return 0
        print(f"{len(hits)} hit(s) in {path.name}:\n")
        print("\n".join(hits))
        return 0

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # Mode: list the sheets. This is the default because these workbooks have
        # up to 35 sheets and a user rarely knows the sheet name up front.
        if not args.sheet:
            print(f"{path.name}  ({len(workbook.sheetnames)} sheets)\n")
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(
                    f"  {sheet_name:32} "
                    f"{worksheet.max_row:>5} rows x {worksheet.max_column:>3} cols"
                )
            return 0

        # Mode: print one sheet. Sheet names are matched case-insensitively so
        # "maintimeline" finds "mainTimeline".
        actual = next(
            (n for n in workbook.sheetnames if n.lower() == args.sheet.lower()), None
        )
        if actual is None:
            print(
                f"No sheet named {args.sheet!r} in {path.name}. "
                f"Run without --sheet to list them.",
                file=sys.stderr,
            )
            return 1

        rows = read_rows(workbook[actual])
        width = max((len(row) for row in rows), default=0)
        print(f"### {path.name} | sheet {actual} | {len(rows)} rows x {width} cols\n")

        if args.format == "records":
            print_records(rows)
        else:
            print_table(rows)
        return 0
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
